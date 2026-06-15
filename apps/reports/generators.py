"""
apps/reports/generators.py — Report Generators (PDF, CSV, Excel)
"""
import csv
import io
import logging
from datetime import date, timedelta

from django.utils import timezone
from django.db.models import Count

from apps.events.analytics import event_type_label, get_or_build_daily_stats

logger = logging.getLogger(__name__)


# ─────────────────────────────── Helpers ─────────────────────────────────────

def _build_doc(buf, title: str):
    from reportlab.platypus import SimpleDocTemplate
    from reportlab.lib.pagesizes import A4
    return SimpleDocTemplate(buf, pagesize=A4, title=title)


def _header_table_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


# ─────────────────────────────── Attendance ──────────────────────────────────

class AttendanceReportGenerator:
    def generate(self, fmt: str, params: dict) -> tuple[bytes, str]:
        from apps.events.models import RecognitionEvent

        date_from = params.get("date_from", str(timezone.now().date()))
        date_to = params.get("date_to", str(timezone.now().date()))
        dept_id = params.get("department_id")

        qs = RecognitionEvent.objects.filter(
            timestamp__date__gte=date_from,
            timestamp__date__lte=date_to,
            event_type="recognized",
        ).select_related("person", "person__department", "camera").order_by("timestamp")

        if dept_id:
            qs = qs.filter(person__department_id=dept_id)

        if fmt == "csv":
            return self._csv(qs), f"attendance_{date_from}_{date_to}.csv"
        if fmt == "xlsx":
            return self._xlsx(qs), f"attendance_{date_from}_{date_to}.xlsx"
        return self._pdf(qs, date_from, date_to), f"attendance_{date_from}_{date_to}.pdf"

    def _csv(self, events) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["ПІБ", "ID", "Підрозділ", "Камера", "Час", "Впевненість (%)"])
        for e in events:
            writer.writerow([
                e.person.full_name if e.person else "",
                e.person.person_id if e.person else "",
                e.person.department.name if e.person and e.person.department else "",
                e.camera.name,
                e.timestamp.strftime("%d.%m.%Y %H:%M:%S"),
                round(e.confidence or 0, 1),
            ])
        return buf.getvalue().encode("utf-8-sig")

    def _xlsx(self, events) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Відвідуваність"

        headers = ["ПІБ", "ID", "Підрозділ", "Камера", "Час", "Впевненість (%)"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1e3a5f")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for e in events:
            ws.append([
                e.person.full_name if e.person else "",
                e.person.person_id if e.person else "",
                e.person.department.name if e.person and e.person.department else "",
                e.camera.name,
                e.timestamp.replace(tzinfo=None),
                round(e.confidence or 0, 1),
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].auto_size = True

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _pdf(self, events, date_from: str, date_to: str) -> bytes:
        from reportlab.platypus import Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = _build_doc(buf, f"Відвідуваність {date_from}–{date_to}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Звіт відвідуваності: {date_from} — {date_to}", styles["Title"]),
            Spacer(1, 12),
        ]

        data = [["ПІБ", "ID", "Підрозділ", "Камера", "Час", "%"]]
        for e in events:
            data.append([
                e.person.full_name if e.person else "–",
                e.person.person_id if e.person else "–",
                e.person.department.name if e.person and e.person.department else "–",
                e.camera.name,
                e.timestamp.strftime("%d.%m.%Y %H:%M"),
                f"{round(e.confidence or 0, 1)}%",
            ])

        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(_header_table_style())
        story.append(table)
        doc.build(story)
        return buf.getvalue()


# ──────────────────────────── Unknown Persons ─────────────────────────────────

class UnknownPersonsReportGenerator:
    def generate(self, fmt: str, params: dict) -> tuple[bytes, str]:
        from apps.events.models import RecognitionEvent

        date_from = params.get("date_from", str(timezone.now().date()))
        date_to = params.get("date_to", str(timezone.now().date()))

        events = RecognitionEvent.objects.filter(
            timestamp__date__gte=date_from,
            timestamp__date__lte=date_to,
            event_type="unknown",
        ).select_related("camera").order_by("timestamp")

        if fmt == "csv":
            return self._csv(events), f"unknown_persons_{date_from}_{date_to}.csv"
        if fmt == "xlsx":
            return self._xlsx(events), f"unknown_persons_{date_from}_{date_to}.xlsx"
        return self._pdf(events, date_from, date_to), f"unknown_persons_{date_from}_{date_to}.pdf"

    def _csv(self, events) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Камера", "Час", "Тривога", "Liveness score"])
        for e in events:
            writer.writerow([
                e.camera.name,
                e.timestamp.strftime("%d.%m.%Y %H:%M:%S"),
                "Так" if e.is_alert else "Ні",
                round(e.liveness_score or 0, 3),
            ])
        return buf.getvalue().encode("utf-8-sig")

    def _xlsx(self, events) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Невідомі особи"
        ws.append(["Камера", "Час", "Тривога", "Liveness score"])

        fill = PatternFill("solid", fgColor="1e3a5f")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

        for e in events:
            ws.append([
                e.camera.name,
                e.timestamp.replace(tzinfo=None),
                "Так" if e.is_alert else "Ні",
                round(e.liveness_score or 0, 3),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _pdf(self, events, date_from: str, date_to: str) -> bytes:
        from reportlab.platypus import Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = _build_doc(buf, f"Невідомі особи {date_from}–{date_to}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Невідомі особи: {date_from} — {date_to}", styles["Title"]),
            Spacer(1, 12),
        ]

        data = [["Камера", "Час", "Тривога", "Liveness"]]
        for e in events:
            data.append([
                e.camera.name,
                e.timestamp.strftime("%d.%m.%Y %H:%M"),
                "Так" if e.is_alert else "Ні",
                f"{round((e.liveness_score or 0) * 100, 0)}%",
            ])

        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(_header_table_style())
        story.append(table)
        doc.build(story)
        return buf.getvalue()


# ──────────────────────────── Security Audit ─────────────────────────────────

class SecurityAuditReportGenerator:
    def generate(self, fmt: str, params: dict) -> tuple[bytes, str]:
        from apps.security.models import SpoofingAttempt, AuditLog

        date_from = params.get("date_from", str(timezone.now().date()))
        date_to = params.get("date_to", str(timezone.now().date()))

        spoofing = SpoofingAttempt.objects.filter(
            detected_at__date__gte=date_from,
            detected_at__date__lte=date_to,
        ).select_related("camera").order_by("detected_at")

        audit = AuditLog.objects.filter(
            timestamp__date__gte=date_from,
            timestamp__date__lte=date_to,
        ).select_related("user").order_by("timestamp")

        if fmt == "csv":
            return self._csv(spoofing, audit), f"security_audit_{date_from}_{date_to}.csv"
        if fmt == "xlsx":
            return self._xlsx(spoofing, audit), f"security_audit_{date_from}_{date_to}.xlsx"
        return self._pdf(spoofing, audit, date_from, date_to), f"security_audit_{date_from}_{date_to}.pdf"

    def _csv(self, spoofing, audit) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)

        writer.writerow(["=== SPOOFING EVENTS ==="])
        writer.writerow(["Камера", "Тип атаки", "EAR", "IP", "Час"])
        for s in spoofing:
            writer.writerow([
                s.camera.name,
                s.attack_type,
                round(s.ear_value or 0, 4),
                s.ip_address or "",
                s.detected_at.strftime("%d.%m.%Y %H:%M:%S"),
            ])

        writer.writerow([])
        writer.writerow(["=== AUDIT LOG ==="])
        writer.writerow(["Користувач", "Дія", "Ресурс", "IP", "Час"])
        for a in audit:
            writer.writerow([
                a.user.username if a.user else "anonymous",
                a.action,
                a.resource_type,
                a.ip_address or "",
                a.timestamp.strftime("%d.%m.%Y %H:%M:%S"),
            ])

        return buf.getvalue().encode("utf-8-sig")

    def _xlsx(self, spoofing, audit) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        # Sheet 1 — Spoofing
        ws1 = wb.active
        ws1.title = "Spoofing"
        ws1.append(["Камера", "Тип атаки", "EAR", "IP", "Час"])
        fill = PatternFill("solid", fgColor="8b1a1a")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws1[1]:
            cell.fill = fill
            cell.font = font
        for s in spoofing:
            ws1.append([
                s.camera.name,
                s.attack_type,
                round(s.ear_value or 0, 4),
                s.ip_address or "",
                s.detected_at.replace(tzinfo=None),
            ])

        # Sheet 2 — Audit Log
        ws2 = wb.create_sheet("Audit Log")
        ws2.append(["Користувач", "Дія", "Ресурс", "IP", "Час"])
        fill2 = PatternFill("solid", fgColor="1e3a5f")
        font2 = Font(color="FFFFFF", bold=True)
        for cell in ws2[1]:
            cell.fill = fill2
            cell.font = font2
        for a in audit:
            ws2.append([
                a.user.username if a.user else "anonymous",
                a.action,
                a.resource_type,
                a.ip_address or "",
                a.timestamp.replace(tzinfo=None),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _pdf(self, spoofing, audit, date_from: str, date_to: str) -> bytes:
        from reportlab.platypus import Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import TableStyle

        buf = io.BytesIO()
        doc = _build_doc(buf, f"Security Audit {date_from}–{date_to}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Аудит безпеки: {date_from} — {date_to}", styles["Title"]),
            Spacer(1, 12),
            Paragraph("Spoofing-атаки", styles["Heading2"]),
            Spacer(1, 6),
        ]

        sp_data = [["Камера", "Тип атаки", "EAR", "IP", "Час"]]
        for s in spoofing:
            sp_data.append([
                s.camera.name,
                s.attack_type,
                str(round(s.ear_value or 0, 4)),
                s.ip_address or "–",
                s.detected_at.strftime("%d.%m.%Y %H:%M"),
            ])

        sp_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8b1a1a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff0f0")]),
        ])
        sp_table = Table(sp_data, repeatRows=1, hAlign="LEFT")
        sp_table.setStyle(sp_style)
        story.append(sp_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph("Аудит дій", styles["Heading2"]))
        story.append(Spacer(1, 6))

        al_data = [["Користувач", "Дія", "Ресурс", "IP", "Час"]]
        for a in audit:
            al_data.append([
                a.user.username if a.user else "anonymous",
                a.action[:60],
                a.resource_type or "–",
                a.ip_address or "–",
                a.timestamp.strftime("%d.%m.%Y %H:%M"),
            ])

        al_table = Table(al_data, repeatRows=1, hAlign="LEFT")
        al_table.setStyle(_header_table_style())
        story.append(al_table)

        doc.build(story)
        return buf.getvalue()


# ────────────────────────────── Daily Summary ─────────────────────────────────

class DailySummaryReportGenerator:
    def generate(self, fmt: str, params: dict) -> tuple[bytes, str]:
        from apps.events.models import RecognitionEvent

        target_date_str = params.get("date_from", str(timezone.now().date()))
        target_date = date.fromisoformat(target_date_str)

        week_stats = [
            get_or_build_daily_stats(target_date - timedelta(days=offset))
            for offset in range(6, -1, -1)
        ]

        # Today's events summary
        events = RecognitionEvent.objects.filter(
            timestamp__date=target_date,
        ).select_related("camera").order_by("camera__name")

        if fmt == "csv":
            return self._csv(week_stats, events, target_date_str), f"daily_summary_{target_date_str}.csv"
        if fmt == "xlsx":
            return self._xlsx(week_stats, events, target_date_str), f"daily_summary_{target_date_str}.xlsx"
        return self._pdf(week_stats, events, target_date_str), f"daily_summary_{target_date_str}.pdf"

    def _csv(self, week_stats, events, target_date: str) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)

        writer.writerow([f"=== ДЕННИЙ ПІДСУМОК: {target_date} ==="])
        writer.writerow(["Дата", "Всього", "Розпізнано", "Невідомі", "Spoofing", "Унікальних осіб"])
        for s in week_stats:
            writer.writerow([
                str(s["date"]), s["total_events"], s["recognized"],
                s["unknown"], s["spoofing_attempts"], s["unique_persons"],
            ])

        writer.writerow([])
        writer.writerow([f"=== АКТИВНІСТЬ ПО КАМЕРАХ ({target_date}) ==="])
        writer.writerow(["Камера", "Тип події", "Кількість"])

        camera_summary = (
            events.values("camera__name", "event_type")
            .annotate(count=Count("id"))
            .order_by("camera__name", "event_type")
        )
        for row in camera_summary:
            writer.writerow([row["camera__name"], event_type_label(row["event_type"]), row["count"]])

        return buf.getvalue().encode("utf-8-sig")

    def _xlsx(self, week_stats, events, target_date: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Тижнева статистика"
        ws1.append(["Дата", "Всього", "Розпізнано", "Невідомі", "Spoofing", "Унік. осіб"])
        fill = PatternFill("solid", fgColor="1e3a5f")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws1[1]:
            cell.fill = fill
            cell.font = font
        for s in week_stats:
            ws1.append([str(s["date"]), s["total_events"], s["recognized"], s["unknown"], s["spoofing_attempts"], s["unique_persons"]])

        ws2 = wb.create_sheet("Камери")
        ws2.append(["Камера", "Тип події", "Кількість"])
        for cell in ws2[1]:
            cell.fill = fill
            cell.font = font

        camera_summary = (
            events.values("camera__name", "event_type")
            .annotate(count=Count("id"))
            .order_by("camera__name", "event_type")
        )
        for row in camera_summary:
            ws2.append([row["camera__name"], event_type_label(row["event_type"]), row["count"]])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _pdf(self, week_stats, events, target_date: str) -> bytes:
        from reportlab.platypus import Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = _build_doc(buf, f"Денний підсумок {target_date}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Денний підсумок: {target_date}", styles["Title"]),
            Spacer(1, 12),
            Paragraph("Статистика за 7 днів", styles["Heading2"]),
            Spacer(1, 6),
        ]

        data = [["Дата", "Всього", "Розпізнано", "Невідомі", "Spoofing", "Унік."]]
        for s in week_stats:
            data.append([
                str(s["date"]), str(s["total_events"]), str(s["recognized"]),
                str(s["unknown"]), str(s["spoofing_attempts"]), str(s["unique_persons"]),
            ])
        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(_header_table_style())
        story.append(table)
        story.append(Spacer(1, 20))

        story.append(Paragraph("Активність по камерах", styles["Heading2"]))
        story.append(Spacer(1, 6))

        camera_summary = list(
            events.values("camera__name", "event_type")
            .annotate(count=Count("id"))
            .order_by("camera__name", "event_type")
        )
        cam_data = [["Камера", "Тип події", "Кількість"]]
        for row in camera_summary:
            cam_data.append([row["camera__name"], event_type_label(row["event_type"]), str(row["count"])])

        cam_table = Table(cam_data, repeatRows=1, hAlign="LEFT")
        cam_table.setStyle(_header_table_style())
        story.append(cam_table)

        doc.build(story)
        return buf.getvalue()


# ────────────────────────────────── Custom ───────────────────────────────────

class CustomReportGenerator:
    """Flexible generator: all events in the date range, all types."""

    def generate(self, fmt: str, params: dict) -> tuple[bytes, str]:
        from apps.events.models import RecognitionEvent

        date_from = params.get("date_from", str(timezone.now().date()))
        date_to = params.get("date_to", str(timezone.now().date()))
        event_type = params.get("event_type")

        qs = RecognitionEvent.objects.filter(
            timestamp__date__gte=date_from,
            timestamp__date__lte=date_to,
        ).select_related("person", "camera").order_by("timestamp")

        if event_type:
            qs = qs.filter(event_type=event_type)

        if fmt == "csv":
            return self._csv(qs), f"custom_{date_from}_{date_to}.csv"
        if fmt == "xlsx":
            return self._xlsx(qs), f"custom_{date_from}_{date_to}.xlsx"
        return self._pdf(qs, date_from, date_to), f"custom_{date_from}_{date_to}.pdf"

    def _csv(self, events) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Тип", "Особа", "ID особи", "Камера", "Впевненість", "Liveness", "Тривога", "Час"])
        for e in events:
            writer.writerow([
                e.get_event_type_display(),
                e.person.full_name if e.person else "Невідома особа",
                e.person.person_id if e.person else "",
                e.camera.name,
                round(e.confidence or 0, 1),
                round(e.liveness_score or 0, 3),
                "Так" if e.is_alert else "Ні",
                e.timestamp.strftime("%d.%m.%Y %H:%M:%S"),
            ])
        return buf.getvalue().encode("utf-8-sig")

    def _xlsx(self, events) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Звіт"
        ws.append(["Тип", "Особа", "ID особи", "Камера", "Впевненість", "Liveness", "Тривога", "Час"])
        fill = PatternFill("solid", fgColor="1e3a5f")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

        for e in events:
            ws.append([
                e.get_event_type_display(),
                e.person.full_name if e.person else "Невідома особа",
                e.person.person_id if e.person else "",
                e.camera.name,
                round(e.confidence or 0, 1),
                round(e.liveness_score or 0, 3),
                "Так" if e.is_alert else "Ні",
                e.timestamp.replace(tzinfo=None),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _pdf(self, events, date_from: str, date_to: str) -> bytes:
        from reportlab.platypus import Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = _build_doc(buf, f"Власний звіт {date_from}–{date_to}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Власний звіт: {date_from} — {date_to}", styles["Title"]),
            Spacer(1, 12),
        ]

        data = [["Тип", "Особа", "Камера", "Впевн.", "Тривога", "Час"]]
        for e in events:
            data.append([
                e.get_event_type_display(),
                e.person.full_name if e.person else "Невідома особа",
                e.camera.name,
                f"{round(e.confidence or 0, 1)}%",
                "Так" if e.is_alert else "Ні",
                e.timestamp.strftime("%d.%m.%Y %H:%M"),
            ])

        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(_header_table_style())
        story.append(table)
        doc.build(story)
        return buf.getvalue()


# ─────────────────────────────── Factory ─────────────────────────────────────

class ReportGeneratorFactory:
    _generators = {
        "attendance":      AttendanceReportGenerator,
        "unknown_persons": UnknownPersonsReportGenerator,
        "security_audit":  SecurityAuditReportGenerator,
        "daily_summary":   DailySummaryReportGenerator,
        "custom":          CustomReportGenerator,
    }

    @classmethod
    def get(cls, report_type: str):
        GeneratorClass = cls._generators.get(report_type)
        if not GeneratorClass:
            raise ValueError(
                f"Unknown report type: '{report_type}'. "
                f"Valid types: {list(cls._generators)}"
            )
        return GeneratorClass()
